import base64
import io
from datetime import date, datetime, time, timedelta
from decimal import Decimal

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader
from reportlab.pdfgen import canvas

from core.models import (
    Contract,
    ContractTemplate,
    Shift,
    TimeEntry,
    WorkingTimeAccountRecord,
)
from core.native_cutover import sync_working_time
from core.working_time import update_record


def _pdf_bytes(text='A+ Payroll Test'):
    output = io.BytesIO()
    pdf = canvas.Canvas(output)
    pdf.drawString(50, 800, text)
    pdf.save()
    return output.getvalue()


def _signature_data_url():
    output = io.BytesIO()
    image = Image.new('RGB', (320, 90), 'white')
    image.save(output, format='PNG')
    return 'data:image/png;base64,' + base64.b64encode(output.getvalue()).decode('ascii')


@pytest.mark.django_db
def test_working_time_keeps_negative_ist_soll_balance_and_exports_it(
    auth_admin, worker_user, company, location, position
):
    worker = worker_user.worker_profile
    worker.monthly_hours = Decimal('10.00')
    worker.tariff_hourly_rate = Decimal('15.50')
    worker.save(update_fields=['monthly_hours', 'tariff_hourly_rate', 'updated_at'])

    today = timezone.localdate()
    clock_in = timezone.make_aware(
        datetime.combine(today, time(8, 0)), timezone.get_current_timezone()
    )
    shift = Shift.objects.create(
        client=company,
        location=location,
        position=position,
        worker=worker,
        starts_at=clock_in,
        ends_at=clock_in + timedelta(hours=8),
        break_minutes=0,
        status=Shift.Status.CONFIRMED,
    )
    TimeEntry.objects.create(
        worker=worker,
        shift=shift,
        clock_in=clock_in,
        clock_out=clock_in + timedelta(hours=8),
        approved=True,
    )

    sync_working_time(today, today)
    record = WorkingTimeAccountRecord.objects.get(
        worker=worker, year_month=today.replace(day=1)
    )
    assert record.ist_hours == Decimal('8.00')
    assert record.soll_hours == Decimal('10.00')
    assert record.difference_hours == Decimal('-2.00')
    assert record.saldo_cumulative == Decimal('-2.00')
    assert record.hourly_rate == Decimal('15.50')
    assert record.gross_amount == Decimal('124.00')

    csv_response = auth_admin.get(
        f'/api/working-time/export/csv/?worker={worker.id}'
    )
    assert csv_response.status_code == 200
    csv_text = csv_response.content.decode('utf-8-sig')
    assert '-2.00' in csv_text
    assert '124.00' in csv_text

    xlsx_response = auth_admin.get(
        f'/api/working-time/export/xlsx/?worker={worker.id}'
    )
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.content), data_only=True)
    row = list(workbook['Arbeitszeitkonto'].iter_rows(values_only=True))[1]
    assert row[2] == 8
    assert row[3] == 10
    assert row[4] == -2
    assert row[8] == -2
    assert row[10] == 124

    pdf_response = auth_admin.get(f'/api/working-time/pdf/{worker.id}/')
    assert pdf_response.status_code == 200
    assert pdf_response.content.startswith(b'%PDF-')


@pytest.mark.django_db
def test_payout_and_manual_adjustment_recalculate_following_month_signed(worker_user):
    worker = worker_user.worker_profile
    first = WorkingTimeAccountRecord.objects.create(
        worker=worker,
        year_month=date(2026, 1, 1),
        ist_hours=Decimal('85.00'),
        soll_hours=Decimal('80.00'),
        difference_hours=Decimal('5.00'),
        saldo_cumulative=Decimal('5.00'),
        hourly_rate=Decimal('15.00'),
        gross_amount=Decimal('1275.00'),
    )
    second = WorkingTimeAccountRecord.objects.create(
        worker=worker,
        year_month=date(2026, 2, 1),
        ist_hours=Decimal('78.00'),
        soll_hours=Decimal('80.00'),
        difference_hours=Decimal('-2.00'),
        carryover_previous=Decimal('5.00'),
        saldo_cumulative=Decimal('3.00'),
        hourly_rate=Decimal('15.00'),
        gross_amount=Decimal('1170.00'),
    )

    update_record(first, paid_hours='4.00', manual_adjustment='0.00')
    first.refresh_from_db()
    second.refresh_from_db()

    assert first.paid_hours == Decimal('4.00')
    assert first.saldo_cumulative == Decimal('1.00')
    assert second.carryover_previous == Decimal('1.00')
    assert second.saldo_cumulative == Decimal('-1.00')


@pytest.mark.django_db
def test_drawn_two_party_signatures_are_hashed_and_stamped_into_final_pdf(
    api_client, worker_user, admin_user
):
    template = ContractTemplate.objects.create(
        name='Drawn Signature Regression',
        slug='drawn-signature-regression',
        kind=ContractTemplate.Kind.EMPLOYMENT,
        source_format=ContractTemplate.SourceFormat.HTML,
        html_template='<h1>Vertrag zur Signatur</h1>',
        schema={'fields': [], 'signature_roles': ['employee', 'employer']},
    )
    contract = Contract.objects.create(
        template=template,
        worker=worker_user.worker_profile,
        title='Drawn Signature Test',
        status=Contract.Status.READY,
        created_by=admin_user,
    )

    signature = _signature_data_url()
    api_client.force_authenticate(worker_user)
    worker_response = api_client.post(
        f'/api/contracts/{contract.id}/sign/',
        {'name': 'Anna Becker', 'signature': signature},
        format='json',
    )
    assert worker_response.status_code == 200

    api_client.force_authenticate(admin_user)
    employer_response = api_client.post(
        f'/api/contracts/{contract.id}/sign/',
        {'name': 'A+ Solution GmbH', 'signature': signature},
        format='json',
    )
    assert employer_response.status_code == 200

    contract.refresh_from_db()
    signatures = list(contract.signatures.order_by('signed_at'))
    assert contract.status == Contract.Status.SIGNED
    assert contract.pdf
    assert len(signatures) == 2
    assert all(len(item.signature_hash) == 64 for item in signatures)
    assert len(contract.signature_hash) == 64
    assert all(item.signature_data.startswith('data:image/png;base64,') for item in signatures)

    contract.pdf.open('rb')
    try:
        pdf_bytes = contract.pdf.read()
    finally:
        contract.pdf.close()
    assert pdf_bytes.startswith(b'%PDF-')
    reader = PdfReader(io.BytesIO(pdf_bytes))
    assert reader.pages
    final_text = reader.pages[-1].extract_text() or ''
    assert 'Anna Becker' in final_text
    assert 'A+ Solution GmbH' in final_text
    assert 'Mitarbeiter' in final_text
    assert 'Arbeitgeber' in final_text


@pytest.mark.django_db
def test_payroll_upload_roundtrip_accepts_real_pdf_and_rejects_fake_files(
    auth_admin, auth_worker, worker_user
):
    worker = worker_user.worker_profile
    valid = SimpleUploadedFile(
        'lohnabrechnung.pdf',
        _pdf_bytes(),
        content_type='application/pdf',
    )
    created = auth_admin.post(
        '/api/payroll/',
        {
            'worker': str(worker.id),
            'period': '2026-01-01',
            'gross_amount': '1240.50',
            'net_amount': '998.10',
            'document': valid,
        },
        format='multipart',
    )
    assert created.status_code == 201
    assert str(created.data['worker']) == str(worker.id)
    assert created.data['gross_amount'] == '1240.50'
    assert created.data['net_amount'] == '998.10'

    worker_list = auth_worker.get('/api/payroll/')
    assert worker_list.status_code == 200
    rows = worker_list.data.get('results', worker_list.data)
    assert len(rows) == 1
    assert str(rows[0]['worker']) == str(worker.id)

    fake_pdf = SimpleUploadedFile(
        'fake.pdf',
        b'not actually a pdf',
        content_type='application/pdf',
    )
    rejected_content = auth_admin.post(
        '/api/payroll/',
        {
            'worker': str(worker.id),
            'period': '2026-02-01',
            'document': fake_pdf,
        },
        format='multipart',
    )
    assert rejected_content.status_code == 400
    assert 'document' in rejected_content.data

    wrong_extension = SimpleUploadedFile(
        'lohnabrechnung.txt',
        _pdf_bytes(),
        content_type='text/plain',
    )
    rejected_type = auth_admin.post(
        '/api/payroll/',
        {
            'worker': str(worker.id),
            'period': '2026-03-01',
            'document': wrong_extension,
        },
        format='multipart',
    )
    assert rejected_type.status_code == 400
    assert 'document' in rejected_type.data
