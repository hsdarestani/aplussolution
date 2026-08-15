import csv
import io
import json
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_CEILING

from django.db.models import Q
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework import serializers, viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from .forecast_models import ForecastDayBudget, ForecastPositionRequirement, ForecastUnitDay, ForecastUnitDefinition
from .models import Shift, User
from .permissions import IsAdminOrManager
from .services import audit
from .shift_slots import ShiftSlot


class ForecastDayBudgetSerializer(serializers.ModelSerializer):
    schedule_name = serializers.CharField(source='schedule.name', read_only=True)

    class Meta:
        model = ForecastDayBudget
        fields = '__all__'


class ForecastPositionRequirementSerializer(serializers.ModelSerializer):
    position_name = serializers.CharField(source='position.name', read_only=True)

    class Meta:
        model = ForecastPositionRequirement
        exclude = ['definition']


class ForecastUnitDefinitionSerializer(serializers.ModelSerializer):
    schedule_name = serializers.CharField(source='schedule.name', read_only=True)
    requirements = ForecastPositionRequirementSerializer(many=True, required=False)

    class Meta:
        model = ForecastUnitDefinition
        fields = '__all__'

    def create(self, validated_data):
        requirements = validated_data.pop('requirements', [])
        obj = ForecastUnitDefinition.objects.create(**validated_data)
        for item in requirements:
            ForecastPositionRequirement.objects.create(definition=obj, **item)
        return obj

    def update(self, instance, validated_data):
        requirements = validated_data.pop('requirements', None)
        instance = super().update(instance, validated_data)
        if requirements is not None:
            instance.requirements.all().delete()
            for item in requirements:
                ForecastPositionRequirement.objects.create(definition=instance, **item)
        return instance


class ForecastUnitDaySerializer(serializers.ModelSerializer):
    definition_name = serializers.CharField(source='definition.name', read_only=True)
    unit_label = serializers.CharField(source='definition.unit_label', read_only=True)

    class Meta:
        model = ForecastUnitDay
        fields = '__all__'


class ManagerForecastViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAdminOrManager]

    def perform_create(self, serializer):
        obj = serializer.save()
        audit(self.request, f'forecast.{obj.__class__.__name__.lower()}.created', obj)

    def perform_update(self, serializer):
        obj = serializer.save()
        audit(self.request, f'forecast.{obj.__class__.__name__.lower()}.updated', obj)


class ForecastDayBudgetViewSet(ManagerForecastViewSet):
    queryset = ForecastDayBudget.objects.select_related('schedule').all()
    serializer_class = ForecastDayBudgetSerializer
    filterset_fields = ['schedule', 'date']
    ordering_fields = ['date']


class ForecastUnitDefinitionViewSet(ManagerForecastViewSet):
    queryset = ForecastUnitDefinition.objects.select_related('schedule').prefetch_related('requirements__position').all()
    serializer_class = ForecastUnitDefinitionSerializer
    filterset_fields = ['schedule', 'mode', 'active']


class ForecastUnitDayViewSet(ManagerForecastViewSet):
    queryset = ForecastUnitDay.objects.select_related('definition', 'definition__schedule').all()
    serializer_class = ForecastUnitDaySerializer
    filterset_fields = ['definition', 'date']
    ordering_fields = ['date']


def _aware(day):
    return timezone.make_aware(datetime.combine(day, datetime.min.time()), timezone.get_current_timezone())


def _shift_minutes(shift):
    return max(0, int((shift.ends_at - shift.starts_at).total_seconds() // 60) - int(shift.break_minutes or 0))


def _schedule_shifts(schedule, start, end):
    locations = list(schedule.locations.values_list('id', flat=True))
    qs = Shift.objects.filter(starts_at__lt=_aware(end), ends_at__gte=_aware(start)).exclude(status=Shift.Status.CANCELLED)
    if locations:
        qs = qs.filter(location_id__in=locations)
    return qs.select_related('position', 'location').prefetch_related('slots__worker')


def _money(value):
    return str(Decimal(value or 0).quantize(Decimal('0.01')))


@api_view(['GET'])
@permission_classes([IsAdminOrManager])
def forecast_summary(request):
    from .scheduling_models import ScheduleGroup

    schedule = ScheduleGroup.objects.prefetch_related('locations').filter(pk=request.GET.get('schedule'), active=True).first()
    if not schedule:
        return Response({'detail': 'Gültiger Dienstplan ist erforderlich.'}, status=400)
    start = parse_date(str(request.GET.get('start') or '')) or timezone.localdate()
    end = parse_date(str(request.GET.get('end') or '')) or (start + timedelta(days=7))
    if end <= start or (end - start).days > 92:
        return Response({'detail': 'Forecast-Zeitraum muss zwischen 1 und 92 Tagen liegen.'}, status=400)

    shifts = list(_schedule_shifts(schedule, start, end))
    budgets = {row.date: row for row in ForecastDayBudget.objects.filter(schedule=schedule, date__gte=start, date__lt=end)}
    definitions = list(ForecastUnitDefinition.objects.filter(schedule=schedule, active=True).prefetch_related('requirements__position', 'days'))
    unit_days = {(row.definition_id, row.date): row for row in ForecastUnitDay.objects.filter(definition__in=definitions, date__gte=start, date__lt=end)}
    days = []
    cursor = start
    while cursor < end:
        day_start, day_end = _aware(cursor), _aware(cursor + timedelta(days=1))
        day_shifts = [s for s in shifts if s.starts_at < day_end and s.ends_at >= day_start]
        combined_minutes = sum(_shift_minutes(s) * max(1, int(s.required_count or 1)) for s in day_shifts)
        assigned_minutes = 0
        labor_cost = Decimal('0')
        position_stats = {}
        for shift in day_shifts:
            minutes = _shift_minutes(shift)
            claimed = [slot for slot in shift.slots.all() if slot.status == ShiftSlot.Status.CLAIMED and slot.worker_id]
            assigned_minutes += minutes * len(claimed)
            stat = position_stats.setdefault(str(shift.position_id), {'position': str(shift.position_id), 'position_name': shift.position.name, 'scheduled_shifts': 0, 'scheduled_hours': Decimal('0')})
            stat['scheduled_shifts'] += max(1, int(shift.required_count or 1))
            stat['scheduled_hours'] += (Decimal(minutes) / Decimal(60)) * max(1, int(shift.required_count or 1))
            for slot in claimed:
                rate = Decimal(slot.worker.tariff_hourly_rate or 0) + Decimal(slot.worker.extra_allowance or 0)
                labor_cost += (Decimal(minutes) / Decimal(60)) * rate

        budget = budgets.get(cursor)
        sales = Decimal((budget.actual_sales if cursor < timezone.localdate() and budget and budget.actual_sales is not None else budget.sales_budget) if budget else 0)
        labor_target = Decimal(budget.labor_percent_target if budget else 0)
        hours_budget = Decimal(budget.hours_budget if budget else 0)
        labor_budget = sales * labor_target / Decimal(100) if labor_target else Decimal(0)
        assigned_percent = (labor_cost / sales * Decimal(100)) if sales else Decimal(0)
        custom = []
        for definition in definitions:
            unit_day = unit_days.get((definition.id, cursor))
            projected = Decimal(unit_day.actual_units if cursor < timezone.localdate() and unit_day and unit_day.actual_units is not None else unit_day.projected_units if unit_day else 0)
            requirements = []
            for requirement in definition.requirements.all():
                raw_required = projected / Decimal(requirement.units_basis) * Decimal(requirement.required_value) if requirement.units_basis else Decimal(0)
                required = raw_required.quantize(Decimal('1'), rounding=ROUND_CEILING) if definition.mode == ForecastUnitDefinition.Mode.SHIFTS else raw_required.quantize(Decimal('0.01'))
                stat = position_stats.get(str(requirement.position_id), {'scheduled_shifts': 0, 'scheduled_hours': Decimal(0)})
                scheduled = Decimal(stat['scheduled_shifts']) if definition.mode == ForecastUnitDefinition.Mode.SHIFTS else Decimal(stat['scheduled_hours'])
                requirements.append({'position': str(requirement.position_id), 'position_name': requirement.position.name, 'required': str(required), 'scheduled': str(scheduled.quantize(Decimal('0.01'))), 'variance': str((scheduled-required).quantize(Decimal('0.01')))})
            custom.append({'definition': str(definition.id), 'name': definition.name, 'unit_label': definition.unit_label, 'mode': definition.mode, 'projected_units': str(projected), 'requirements': requirements})

        combined_hours = Decimal(combined_minutes) / Decimal(60)
        assigned_hours = Decimal(assigned_minutes) / Decimal(60)
        days.append({
            'date': cursor.isoformat(),
            'sales_budget': _money(sales),
            'labor_percent_target': _money(labor_target),
            'labor_budget': _money(labor_budget),
            'assigned_labor_cost': _money(labor_cost),
            'assigned_labor_percent': _money(assigned_percent),
            'labor_variance': _money(labor_budget-labor_cost),
            'hours_budget': _money(hours_budget),
            'combined_hours': _money(combined_hours),
            'assigned_hours': _money(assigned_hours),
            'hours_variance': _money(hours_budget-combined_hours),
            'custom_units': custom,
        })
        cursor += timedelta(days=1)
    return Response({'schedule': str(schedule.id), 'schedule_name': schedule.name, 'start': start, 'end': end, 'days': days})


def _read_upload(upload):
    if upload.size > 8 * 1024 * 1024:
        raise ValueError('Datei ist größer als 8 MB.')
    name = upload.name.lower()
    if name.endswith('.csv'):
        text = upload.read().decode('utf-8-sig')
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=',;\t')
        except csv.Error:
            dialect = csv.excel
        return list(csv.DictReader(io.StringIO(text), dialect=dialect))
    if name.endswith('.xlsx'):
        from openpyxl import load_workbook
        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    elif name.endswith('.xls'):
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError('XLS-Unterstützung ist nicht installiert.') from exc
        book = xlrd.open_workbook(file_contents=upload.read())
        sheet = book.sheet_by_index(0)
        rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    else:
        raise ValueError('Nur CSV, XLS und XLSX werden unterstützt.')
    if not rows:
        return []
    headers = [str(value or '').strip() for value in rows[0]]
    return [{headers[index]: value for index, value in enumerate(row) if index < len(headers)} for row in rows[1:]]


@api_view(['POST'])
@permission_classes([IsAdminOrManager])
def forecast_import_preview(request):
    upload = request.FILES.get('file')
    if not upload:
        return Response({'detail': 'Forecast-Datei fehlt.'}, status=400)
    try:
        rows = _read_upload(upload)
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=400)
    headers = list(rows[0].keys()) if rows else []
    return Response({'headers': headers, 'preview': rows[:10], 'row_count': len(rows)})


def _decimal(value):
    if value in (None, ''):
        return Decimal(0)
    cleaned = str(value).strip().replace('€', '').replace('$', '').replace(' ', '')
    if ',' in cleaned and '.' not in cleaned:
        cleaned = cleaned.replace(',', '.')
    elif ',' in cleaned and '.' in cleaned:
        cleaned = cleaned.replace(',', '')
    return Decimal(cleaned)


@api_view(['POST'])
@permission_classes([IsAdminOrManager])
def forecast_import_apply(request):
    from .scheduling_models import ScheduleGroup

    upload = request.FILES.get('file')
    schedule = ScheduleGroup.objects.filter(pk=request.data.get('schedule'), active=True).first()
    if not upload or not schedule:
        return Response({'detail': 'Forecast-Datei und Dienstplan sind erforderlich.'}, status=400)
    try:
        mapping = json.loads(request.data.get('mapping') or '{}') if isinstance(request.data.get('mapping'), str) else (request.data.get('mapping') or {})
        rows = _read_upload(upload)
    except (ValueError, json.JSONDecodeError) as exc:
        return Response({'detail': str(exc)}, status=400)
    date_column = mapping.get('date')
    if not date_column:
        return Response({'detail': 'Datums-Spalte muss zugeordnet werden.'}, status=400)
    created = updated = 0
    errors = []
    unit_targets = {key.split(':', 1)[1]: source for key, source in mapping.items() if key.startswith('unit:') and source}
    definitions = {str(item.id): item for item in ForecastUnitDefinition.objects.filter(schedule=schedule, active=True)}
    for index, row in enumerate(rows, start=2):
        try:
            raw_date = row.get(date_column)
            if isinstance(raw_date, datetime):
                day = raw_date.date()
            else:
                day = parse_date(str(raw_date).strip())
                if not day:
                    for fmt in ('%d.%m.%Y', '%m/%d/%Y', '%d/%m/%Y'):
                        try:
                            day = datetime.strptime(str(raw_date).strip(), fmt).date(); break
                        except ValueError:
                            continue
            if not day:
                raise ValueError('Ungültiges Datum')
            defaults = {
                'sales_budget': _decimal(row.get(mapping.get('sales_budget'))) if mapping.get('sales_budget') else 0,
                'hours_budget': _decimal(row.get(mapping.get('hours_budget'))) if mapping.get('hours_budget') else 0,
                'labor_percent_target': _decimal(row.get(mapping.get('labor_percent_target'))) if mapping.get('labor_percent_target') else 0,
            }
            if mapping.get('actual_sales'):
                value = row.get(mapping['actual_sales'])
                defaults['actual_sales'] = _decimal(value) if value not in (None, '') else None
            _, was_created = ForecastDayBudget.objects.update_or_create(schedule=schedule, date=day, defaults=defaults)
            created += int(was_created); updated += int(not was_created)
            for definition_id, source_column in unit_targets.items():
                definition = definitions.get(definition_id)
                if definition:
                    ForecastUnitDay.objects.update_or_create(definition=definition, date=day, defaults={'projected_units': _decimal(row.get(source_column))})
        except Exception as exc:
            errors.append({'line': index, 'error': str(exc)})
    audit(request, 'forecast.imported', schedule, {'created': created, 'updated': updated, 'errors': len(errors)})
    return Response({'created': created, 'updated': updated, 'errors': errors})
